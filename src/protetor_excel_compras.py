"""
Monitor inteligente de proteção de Excel
Detecta tentativas de alteração e restaura valores originais
Processa compras quando célula de ação é modificada
"""
import os
import json
import threading
import time
import logging
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - PROTETOR_EXCEL - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ProtectorExcelCompras:
    """Monitora e protege as abas de compra contra edição não autorizada"""
    
    def __init__(self, pasta_equipes: str = "data/equipes"):
        """
        Inicializa o protetor
        
        Args:
            pasta_equipes: Pasta onde estão os Excels
        """
        self.pasta_equipes = Path(pasta_equipes)
        self.rodando = False
        self.thread_monitor = None
        
        # Cache de valores originais
        self.valores_originais = {}
        
        logger.info("🛡️  Protetor de Excel de Compras inicializado")
    
    def iniciar(self):
        """Inicia o monitoramento"""
        if self.rodando:
            return
        
        self.rodando = True
        self.thread_monitor = threading.Thread(
            target=self._monitor_loop,
            daemon=True,
            name="ProtectorExcelCompras"
        )
        self.thread_monitor.start()
        logger.info("🛡️  Protetor iniciado - monitorando edições não autorizadas")
    
    def parar(self):
        """Para o monitoramento"""
        self.rodando = False
        if self.thread_monitor:
            self.thread_monitor.join(timeout=5)
        logger.info("🛑 Protetor parado")
    
    def _monitor_loop(self):
        """Loop de monitoramento"""
        while self.rodando:
            try:
                self._verificar_excels()
                time.sleep(3)  # Verificar a cada 3 segundos
            except Exception as e:
                logger.error(f"Erro no monitoramento: {e}")
                time.sleep(1)
    
    def _verificar_excels(self):
        """Verifica todos os Excels para detecção de alterações"""
        if not self.pasta_equipes.exists():
            return
        
        for arquivo_xlsx in self.pasta_equipes.glob("*.xlsx"):
            try:
                self._verificar_arquivo(arquivo_xlsx)
            except Exception as e:
                logger.debug(f"Erro ao verificar {arquivo_xlsx.name}: {e}")
    
    def _verificar_arquivo(self, caminho_arquivo):
        """Verifica um arquivo específico para detecção de cliques/alterações"""
        try:
            wb = load_workbook(caminho_arquivo, data_only=False)
            
            # Verificar abas de loja
            for sheet_name in ['Loja Carros', 'Loja Peças']:
                if sheet_name not in wb.sheetnames:
                    continue
                
                ws = wb[sheet_name]
                col_acao = 'G' if sheet_name == 'Loja Carros' else 'F'
                
                # Verificar coluna de ação
                for row_idx in range(2, ws.max_row + 1):  # Pular header
                    cell = ws[f'{col_acao}{row_idx}']
                    cell_value = cell.value
                    
                    # Se foi alterado para algo diferente de "🛒 COMPRAR"
                    if cell_value and cell_value != "🛒 COMPRAR":
                        # Log e restauração
                        logger.warning(f"⚠️  Tentativa de edição detectada em {caminho_arquivo.name} - {sheet_name} linha {row_idx}")
                        logger.info(f"🔄 Restaurando valor original: '🛒 COMPRAR'")
                        
                        # Restaurar valor original
                        cell.value = "🛒 COMPRAR"
                        wb.save(caminho_arquivo)
            
            wb.close()
        except Exception as e:
            logger.debug(f"Erro ao verificar {caminho_arquivo.name}: {e}")

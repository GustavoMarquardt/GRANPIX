"""
Adiciona botões interativos (Form Controls) ao Excel para compras 100% automáticas
Quando clicado, o botão escreve a solicitação JSON e o sistema processa IMEDIATAMENTE
"""
import os
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def adicionar_botoes_compra(caminho_excel, equipe_id, equipe_nome, loja_carros=None, loja_pecas=None):
    """
    Adiciona IDs de itens nas colunas ocultas para VBA usar
    
    Args:
        caminho_excel: Caminho do arquivo Excel
        equipe_id: ID da equipe
        equipe_nome: Nome da equipe
        loja_carros: Instância de LojaCarros ou lista
        loja_pecas: Instância de LojaPecas ou lista
    """
    try:
        wb = load_workbook(caminho_excel)
        
        # ========== LOJA CARROS ==========
        if 'Loja Carros' in wb.sheetnames and loja_carros:
            ws = wb['Loja Carros']
            
            # Coluna oculta para armazenar IDs (coluna Y)
            col_id = 'Y'
            ws.column_dimensions[col_id].hidden = True
            
            # Obter lista de carros (se é objeto LojaCarros, pega o atributo)
            carros_lista = getattr(loja_carros, 'modelos', loja_carros) if loja_carros else []
            
            for row_idx in range(2, min(50, ws.max_row + 1)):
                marca_cell = ws[f'A{row_idx}']
                modelo_cell = ws[f'B{row_idx}']
                
                if not marca_cell.value or not modelo_cell.value:
                    break
                
                # Procurar carro correspondente
                try:
                    for carro in carros_lista:
                        if str(marca_cell.value).strip().lower() == str(carro.marca).lower() and \
                           str(modelo_cell.value).strip().lower() == str(carro.modelo).lower():
                            # Armazenar ID na coluna oculta
                            ws[f'{col_id}{row_idx}'] = carro.id
                            break
                except:
                    pass  # Continuar se houver erro
        
        # ========== LOJA PEÇAS ==========
        if 'Loja Peças' in wb.sheetnames and loja_pecas:
            ws = wb['Loja Peças']
            
            # Coluna oculta para armazenar IDs (coluna Y)
            col_id = 'Y'
            ws.column_dimensions[col_id].hidden = True
            
            # Obter lista de peças (se é objeto LojaPecas, pega o atributo)
            pecas_lista = getattr(loja_pecas, 'pecas', loja_pecas) if loja_pecas else []
            
            for row_idx in range(2, min(50, ws.max_row + 1)):
                nome_cell = ws[f'A{row_idx}']
                
                if not nome_cell.value:
                    break
                
                # Procurar peça correspondente
                try:
                    for peca in pecas_lista:
                        if str(nome_cell.value).strip().lower() == str(peca.nome).lower():
                            # Armazenar ID na coluna oculta
                            ws[f'{col_id}{row_idx}'] = peca.id
                            break
                except:
                    pass  # Continuar se houver erro
        
        wb.save(caminho_excel)
        return True
        
    except Exception as e:
        # Não mostrar erro durante export - é apenas preparação de botões
        return False


def criar_arquivo_solicitacoes():
    """Cria arquivo vazio de solicitações se não existir"""
    pasta = "data/solicitacoes_compra"
    os.makedirs(pasta, exist_ok=True)
    
    arquivo = os.path.join(pasta, "solicitacoes.json")
    if not os.path.exists(arquivo):
        with open(arquivo, 'w', encoding='utf-8') as f:
            json.dump([], f)
    
    return arquivo


def criar_macro_vba(equipe_id, equipe_nome, tipo_item='peca'):
    """Placeholder para geração de VBA - não usado por enquanto"""
    return ""


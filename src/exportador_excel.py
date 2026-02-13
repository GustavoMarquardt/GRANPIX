"""
Módulo para exportação de dados das equipes para Excel
"""
import os
from typing import List, Optional
from .models import Equipe
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from adicionar_botoes_excel import adicionar_botoes_compra, criar_arquivo_solicitacoes


class ExportadorEquipes:
    """Exporta dados das equipes para arquivos Excel com abas de loja"""
    
    def __init__(self, pasta_saida: Optional[str] = None, usar_onedrive: bool = False, 
                 loja_carros=None, loja_pecas=None):
        """Inicializa o exportador
        
        Args:
            pasta_saida: Pasta onde os arquivos Excel serão salvos
            usar_onedrive: Se True, tenta usar OneDrive se disponível
            loja_carros: Instância de LojaCarros
            loja_pecas: Instância de LojaPecas
        """
        self.loja_carros = loja_carros
        self.loja_pecas = loja_pecas
        
        if usar_onedrive:
            pasta_onedrive = self._detectar_onedrive()
            if pasta_onedrive:
                self.pasta_saida = pasta_onedrive
                self.usando_onedrive = True
                print(f"📁 Usando OneDrive: {self.pasta_saida}")
            else:
                # Fallback para pasta local
                self.pasta_saida = pasta_saida or "data/equipes"
                self.usando_onedrive = False
                print(f"⚠️ OneDrive não detectado, usando: {self.pasta_saida}")
        else:
            self.pasta_saida = pasta_saida or "data/equipes"
            self.usando_onedrive = False
        
        # Criar pasta se não existir
        os.makedirs(self.pasta_saida, exist_ok=True)
    
    def _detectar_onedrive(self) -> Optional[str]:
        """Detecta se OneDrive está disponível e retorna o caminho
        
        Returns:
            Caminho do OneDrive se encontrado, None caso contrário
        """
        # Caminhos possíveis do OneDrive no Windows
        home = str(Path.home())
        onedrive_paths = [
            os.path.join(home, "OneDrive"),  # Padrão
            os.path.join(home, "Documents"),  # Com Documents
        ]
        
        for path in onedrive_paths:
            if os.path.exists(path):
                # Criar estrutura de pasta GRANPIX/equipes
                granpix_path = os.path.join(path, "GRANPIX", "equipes")
                # Criar a pasta se não existir
                os.makedirs(granpix_path, exist_ok=True)
                return granpix_path
        
        return None
    
    def obter_status_onedrive(self) -> str:
        """Retorna o status do OneDrive
        
        Returns:
            String com status e caminho
        """
        if self.usando_onedrive:
            return f"✅ OneDrive ativo: {self.pasta_saida}"
        else:
            return f"📁 Local: {self.pasta_saida}"
    
    def ativar_onedrive(self) -> bool:
        """Ativa a exportação para OneDrive se disponível
        
        Returns:
            True se OneDrive foi ativado, False caso contrário
        """
        pasta_onedrive = self._detectar_onedrive()
        if pasta_onedrive:
            self.pasta_saida = pasta_onedrive
            self.usando_onedrive = True
            print(f"✅ OneDrive ativado: {self.pasta_saida}")
            return True
        else:
            print("❌ OneDrive não detectado no sistema")
            return False
    
    def desativar_onedrive(self):
        """Desativa OneDrive e volta para pasta local"""
        self.pasta_saida = "data/equipes"
        self.usando_onedrive = False
        os.makedirs(self.pasta_saida, exist_ok=True)
        print(f"✅ Voltando para pasta local: {self.pasta_saida}")
    
    def exportar_equipe(self, equipe: Equipe) -> str:
        """Exporta dados de uma equipe para um arquivo Excel com abas de loja
        Se arquivo já existe, apenas atualiza. Se não existe, cria novo.
        
        Args:
            equipe: Equipe a ser exportada
            
        Returns:
            Caminho do arquivo gerado/atualizado
        """
        # Nome base do arquivo (sem timestamp)
        nome_base = f"{equipe.nome.replace(' ', '_').lower()}.xlsx"
        caminho_arquivo = os.path.join(self.pasta_saida, nome_base)
        
        # Verificar se arquivo já existe
        arquivo_existe = os.path.exists(caminho_arquivo)
        
        if arquivo_existe:
            # Atualizar arquivo existente
            return self._atualizar_arquivo_excel(equipe, caminho_arquivo)
        else:
            # Criar novo arquivo
            return self._criar_arquivo_excel(equipe, caminho_arquivo)
    
    def _criar_arquivo_excel(self, equipe: Equipe, caminho_arquivo: str) -> str:
        """Cria um novo arquivo Excel do zero
        
        Args:
            equipe: Equipe a ser exportada
            caminho_arquivo: Caminho onde salvar
            
        Returns:
            Caminho do arquivo criado
        """
        # Criar writer Excel
        with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
            # Aba 1: Informações da Equipe
            self._exportar_info_equipe(equipe, writer)
            
            # Aba 2: Dados do Carro
            self._exportar_dados_carro(equipe, writer)
            
            # Aba 3: Peças Instaladas
            self._exportar_pecas_instaladas(equipe, writer)
            
            # Aba 4: Desgaste das Peças
            self._exportar_desgaste_pecas(equipe, writer)
            
            # Aba 5: Estatísticas
            self._exportar_estatisticas(equipe, writer)
            
            # Aba 6: Loja de Carros (se disponível)
            if self.loja_carros:
                self._exportar_loja_carros(equipe, writer)
            
            # Aba 7: Loja de Peças (se disponível)
            if self.loja_pecas:
                self._exportar_loja_pecas(equipe, writer)
        
        # Proteger as planilhas após criar
        self._proteger_planilhas(caminho_arquivo, equipe.id)
        
        # ⭐ NOVO: Adicionar botões automáticos para compras
        adicionar_botoes_compra(caminho_arquivo, equipe.id, equipe.nome, 
                               self.loja_carros, self.loja_pecas)
        
        # Garantir que arquivo de solicitações existe
        criar_arquivo_solicitacoes()
        
        return caminho_arquivo
    
    def _atualizar_arquivo_excel(self, equipe: Equipe, caminho_arquivo: str) -> str:
        """Atualiza um arquivo Excel existente com novos dados
        
        Args:
            equipe: Equipe com dados atualizados
            caminho_arquivo: Caminho do arquivo a atualizar
            
        Returns:
            Caminho do arquivo atualizado
        """
        try:
            # Carregar workbook existente
            wb = load_workbook(caminho_arquivo)
            
            # Atualizar cada aba existente
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.delete_rows(2, ws.max_row)  # Apagar dados antigos, manter header
            
            # Recriar dados com writer
            with pd.ExcelWriter(caminho_arquivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                # Aba 1: Informações da Equipe
                self._exportar_info_equipe(equipe, writer)
                
                # Aba 2: Dados do Carro
                self._exportar_dados_carro(equipe, writer)
                
                # Aba 3: Peças Instaladas
                self._exportar_pecas_instaladas(equipe, writer)
                
                # Aba 4: Desgaste das Peças
                self._exportar_desgaste_pecas(equipe, writer)
                
                # Aba 5: Estatísticas
                self._exportar_estatisticas(equipe, writer)
                
                # Aba 6: Loja de Carros (se disponível)
                if self.loja_carros and 'Loja Carros' not in wb.sheetnames:
                    self._exportar_loja_carros(equipe, writer)
                
                # Aba 7: Loja de Peças (se disponível)
                if self.loja_pecas and 'Loja Peças' not in wb.sheetnames:
                    self._exportar_loja_pecas(equipe, writer)
            
            # Proteger as planilhas após atualizar
            self._proteger_planilhas(caminho_arquivo, equipe.id)
            
            # ⭐ NOVO: Atualizar botões de compra
            adicionar_botoes_compra(caminho_arquivo, equipe.id, equipe.nome, 
                                   self.loja_carros, self.loja_pecas)
            
            return caminho_arquivo
            
        except Exception as e:
            print(f"⚠️ Erro ao atualizar {caminho_arquivo}: {e}")
            print(f"🔄 Criando novo arquivo...")
            # Se der erro na atualização, cria um novo
            nome_base = os.path.basename(caminho_arquivo)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_novo = nome_base.replace('.xlsx', f'_{timestamp}.xlsx')
            caminho_novo = os.path.join(self.pasta_saida, nome_novo)
            return self._criar_arquivo_excel(equipe, caminho_novo)
    
    def _exportar_info_equipe(self, equipe: Equipe, writer):
        """Exporta informações gerais da equipe (SEM dados de pilotos)"""
        dados = {
            'Propriedade': [
                'Nome da Equipe',
                'ID da Equipe',
                'Doricoins (Dinheiro)',
                'Carro Atual',
                'Data de Exportação'
            ],
            'Valor': [
                equipe.nome,
                equipe.id,
                f"💰 {equipe.doricoins:,.2f}",
                f"{equipe.carro.marca} {equipe.carro.modelo}" if equipe.carro else "N/A",
                datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            ]
        }
        df = pd.DataFrame(dados)
        df.to_excel(writer, sheet_name='Equipe', index=False)
        
        # Formatar coluna de Propriedade
        worksheet = writer.sheets['Equipe']
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 40
    
    def _exportar_dados_carro(self, equipe: Equipe, writer):
        """Exporta dados do carro"""
        if not equipe.carro:
            df = pd.DataFrame({'Aviso': ['Nenhum carro atribuído à equipe']})
            df.to_excel(writer, sheet_name='Carro', index=False)
            return
        
        carro = equipe.carro
        dados = {
            'Propriedade': [
                'Número do Carro',
                'Marca',
                'Modelo',
                'Condição Geral',
                'Batalhas Totais',
                'Vitórias',
                'Derrotas',
                'Empates'
            ],
            'Valor': [
                f"#{carro.numero_carro}",
                carro.marca,
                carro.modelo,
                f"{carro.calcular_condicao_geral():.1f}%",
                carro.batidas_totais,
                carro.vitoria,
                carro.derrotas,
                carro.empates
            ]
        }
        df = pd.DataFrame(dados)
        df.to_excel(writer, sheet_name='Carro', index=False)
        
        worksheet = writer.sheets['Carro']
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 40
    
    def _exportar_pecas_instaladas(self, equipe: Equipe, writer):
        """Exporta peças instaladas no carro"""
        if not equipe.carro or not equipe.carro.pecas_instaladas:
            df = pd.DataFrame({'Aviso': ['Nenhuma peça adicional instalada']})
            df.to_excel(writer, sheet_name='Peças Adicionais', index=False)
            return
        
        dados = {
            'Nome': [],
            'Tipo': [],
            'Preço': [],
            'Durabilidade': []
        }
        
        for peca in equipe.carro.pecas_instaladas:
            dados['Nome'].append(peca.get('nome', 'N/A'))
            dados['Tipo'].append(peca.get('tipo', 'N/A'))
            dados['Preço'].append(f"💰 {peca.get('preco', 0):,.2f}")
            durability = peca.get('durabilidade_atual', 100) / peca.get('durabilidade_maxima', 100) * 100
            dados['Durabilidade'].append(f"{durability:.1f}%")
        
        df = pd.DataFrame(dados)
        df.to_excel(writer, sheet_name='Peças Adicionais', index=False)
        
        worksheet = writer.sheets['Peças Adicionais']
        for col in worksheet.columns:
            worksheet.column_dimensions[col[0].column_letter].width = 25
    
    def _exportar_desgaste_pecas(self, equipe: Equipe, writer):
        """Exporta desgaste das peças principais"""
        if not equipe.carro:
            df = pd.DataFrame({'Aviso': ['Nenhum carro atribuído']})
            df.to_excel(writer, sheet_name='Desgaste Peças', index=False)
            return
        
        carro = equipe.carro
        dados = {
            'Peça': [],
            'Tipo': [],
            'Durabilidade Máxima': [],
            'Durabilidade Atual': [],
            'Percentual': [],
            'Status': []
        }
        
        for peca in carro.get_todas_pecas():
            durability_pct = (peca.durabilidade_atual / peca.durabilidade_maxima) * 100
            
            if durability_pct >= 70:
                status = "🟢 Bom"
            elif durability_pct >= 40:
                status = "🟡 Regular"
            else:
                status = "🔴 Crítico"
            
            dados['Peça'].append(peca.nome)
            dados['Tipo'].append(peca.tipo)
            dados['Durabilidade Máxima'].append(f"{peca.durabilidade_maxima:.1f}%")
            dados['Durabilidade Atual'].append(f"{peca.durabilidade_atual:.1f}%")
            dados['Percentual'].append(f"{durability_pct:.1f}%")
            dados['Status'].append(status)
        
        df = pd.DataFrame(dados)
        df.to_excel(writer, sheet_name='Desgaste Peças', index=False)
        
        worksheet = writer.sheets['Desgaste Peças']
        for col in worksheet.columns:
            worksheet.column_dimensions[col[0].column_letter].width = 25
    
    def _exportar_estatisticas(self, equipe: Equipe, writer):
        """Exporta estatísticas gerais"""
        if not equipe.carro:
            df = pd.DataFrame({'Aviso': ['Nenhum carro para estatísticas']})
            df.to_excel(writer, sheet_name='Estatísticas', index=False)
            return
        
        carro = equipe.carro
        total_batalhas = carro.vitoria + carro.derrotas + carro.empates
        taxa_vitoria = (carro.vitoria / total_batalhas * 100) if total_batalhas > 0 else 0
        
        dados = {
            'Métrica': [
                'Total de Batalhas',
                'Vitórias',
                'Derrotas',
                'Empates',
                'Taxa de Vitória',
                'Coeficiente Quebra (Motor)',
                'Coeficiente Quebra (Câmbio)',
                'Coeficiente Quebra (Kit Ângulo)',
                'Coeficiente Quebra (Suspensão)'
            ],
            'Valor': [
                total_batalhas,
                carro.vitoria,
                carro.derrotas,
                carro.empates,
                f"{taxa_vitoria:.1f}%",
                f"{carro.motor.coeficiente_quebra:.3f}",
                f"{carro.cambio.coeficiente_quebra:.3f}",
                f"{carro.kit_angulo.coeficiente_quebra:.3f}",
                f"{carro.suspensao.coeficiente_quebra:.3f}"
            ]
        }
        
        df = pd.DataFrame(dados)
        df.to_excel(writer, sheet_name='Estatísticas', index=False)
        
        worksheet = writer.sheets['Estatísticas']
        worksheet.column_dimensions['A'].width = 35
        worksheet.column_dimensions['B'].width = 30
    
    def exportar_todas_equipes(self, equipes: List[Equipe]) -> List[str]:
        """Exporta dados de todas as equipes
        
        Args:
            equipes: Lista de equipes a exportar
            
        Returns:
            Lista de caminhos dos arquivos gerados
        """
        arquivos = []
        for equipe in equipes:
            try:
                caminho = self.exportar_equipe(equipe)
                arquivos.append(caminho)
                print(f"✓ Equipe '{equipe.nome}' exportada: {caminho}")
            except Exception as e:
                print(f"✗ Erro ao exportar equipe '{equipe.nome}': {e}")
        
        return arquivos
    
    def exportar_equipe_silencioso(self, equipe: Equipe) -> str:
        """Exporta dados de uma equipe para Excel sem mensagens
        
        Útil para exportação automática após batalhas
        
        Args:
            equipe: Equipe a ser exportada
            
        Returns:
            Caminho do arquivo gerado
        """
        # Nome do arquivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"{equipe.nome.replace(' ', '_').lower()}_{timestamp}.xlsx"
        caminho_arquivo = os.path.join(self.pasta_saida, nome_arquivo)
        
        # Criar writer Excel
        with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
            # Aba 1: Informações da Equipe
            self._exportar_info_equipe(equipe, writer)
            
            # Aba 2: Dados do Carro
            self._exportar_dados_carro(equipe, writer)
            
            # Aba 3: Peças Instaladas
            self._exportar_pecas_instaladas(equipe, writer)
            
            # Aba 4: Desgaste das Peças
            self._exportar_desgaste_pecas(equipe, writer)
            
            # Aba 5: Estatísticas
            self._exportar_estatisticas(equipe, writer)
            
            # Aba 6: Loja de Carros (se disponível)
            if self.loja_carros:
                self._exportar_loja_carros(equipe, writer)
            
            # Aba 7: Loja de Peças (se disponível)
            if self.loja_pecas:
                self._exportar_loja_pecas(equipe, writer)
        
        # Proteger as planilhas após criar
        self._proteger_planilhas(caminho_arquivo, equipe.id)
        
        return caminho_arquivo
    
    def _exportar_loja_carros(self, equipe: Equipe, writer):
        """Exporta catálogo da loja de carros"""
        if not self.loja_carros or not self.loja_carros.modelos:
            df = pd.DataFrame({'Aviso': ['Loja de carros não disponível']})
            df.to_excel(writer, sheet_name='Loja Carros', index=False)
            return
        
        dados = {
            'ID': [],
            'Marca': [],
            'Modelo': [],
            'Classe': [],
            'Preço': [],
            'Descrição': [],
            'Ação': []
        }
        
        for modelo in self.loja_carros.modelos:
            dados['ID'].append(modelo.id)
            dados['Marca'].append(modelo.marca)
            dados['Modelo'].append(modelo.modelo)
            dados['Classe'].append(modelo.classe)
            dados['Preço'].append(f"💰 {modelo.preco:,.2f}")
            dados['Descrição'].append(modelo.descricao)
            dados['Ação'].append("🛒 COMPRAR")
        
        df = pd.DataFrame(dados)
        df.to_excel(writer, sheet_name='Loja Carros', index=False)
        
        worksheet = writer.sheets['Loja Carros']
        # Formatação
        for col_idx, col_name in enumerate(['ID', 'Marca', 'Modelo', 'Classe', 'Preço', 'Descrição', 'Ação'], 1):
            worksheet.column_dimensions[chr(64 + col_idx)].width = 15 if col_name in ['ID', 'Classe', 'Ação'] else 20
        
        # Colorir coluna de ação
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        for row in range(2, len(dados['ID']) + 2):
            worksheet[f'G{row}'].fill = green_fill
    
    def _exportar_loja_pecas(self, equipe: Equipe, writer):
        """Exporta catálogo da loja de peças"""
        if not self.loja_pecas or not self.loja_pecas.pecas:
            df = pd.DataFrame({'Aviso': ['Loja de peças não disponível']})
            df.to_excel(writer, sheet_name='Loja Peças', index=False)
            return
        
        dados = {
            'ID': [],
            'Nome': [],
            'Tipo': [],
            'Preço': [],
            'Descrição': [],
            'Ação': []
        }
        
        for peca in self.loja_pecas.pecas:
            dados['ID'].append(peca.id)
            dados['Nome'].append(peca.nome)
            dados['Tipo'].append(peca.tipo)
            dados['Preço'].append(f"💰 {peca.preco:,.2f}")
            dados['Descrição'].append(peca.descricao)
            dados['Ação'].append("🛒 COMPRAR")
        
        df = pd.DataFrame(dados)
        df.to_excel(writer, sheet_name='Loja Peças', index=False)
        
        worksheet = writer.sheets['Loja Peças']
        # Formatação
        for col_idx, col_name in enumerate(['ID', 'Nome', 'Tipo', 'Preço', 'Descrição', 'Ação'], 1):
            worksheet.column_dimensions[chr(64 + col_idx)].width = 15 if col_name in ['ID', 'Tipo', 'Ação'] else 20
        
        # Colorir coluna de ação
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        for row in range(2, len(dados['ID']) + 2):
            worksheet[f'F{row}'].fill = green_fill
    
    def _proteger_planilhas(self, caminho_arquivo: str, equipe_id: str):
        """Protege abas contra edição
        Abas de loja (Loja Carros, Loja Peças): apenas coluna "Ação" desbloqueada
        Outras abas: totalmente bloqueadas
        
        Args:
            caminho_arquivo: Caminho do arquivo Excel
            equipe_id: ID da equipe (usado como senha)
        """
        try:
            from openpyxl.styles import Protection
            wb = load_workbook(caminho_arquivo)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Definir coluna de ação por aba
                col_acao = None
                if sheet_name == 'Loja Carros':
                    col_acao = 'G'  # Coluna G é "Ação"
                elif sheet_name == 'Loja Peças':
                    col_acao = 'F'  # Coluna F é "Ação"
                
                # Bloquear TODAS as células primeiro
                for row in ws.iter_rows():
                    for cell in row:
                        cell.protection = Protection(locked=True)
                
                # Se for aba de loja, desbloquear APENAS coluna de ação
                if col_acao:
                    for row in ws.iter_rows():
                        cell_acao = row[ord(col_acao) - ord('A')]  # Converter letra para índice
                        cell_acao.protection = Protection(locked=False)
                    print(f"   🔒 {sheet_name}: Protegido (apenas coluna '{col_acao}' desbloqueada para compras)")
                else:
                    # Para outras abas, bloquear tudo
                    print(f"   🔒 {sheet_name}: Totalmente protegido (dados seguros)")
                
                # Proteger a planilha permitindo apenas selecionar
                ws.protection.sheet = True
                ws.protection.enable()
                ws.protection.password = "equipe"
                ws.protection.allowInsertRows = False
                ws.protection.allowDeleteRows = False
                ws.protection.allowInsertColumns = False
                ws.protection.allowDeleteColumns = False
                ws.protection.allowFormatCells = False
                ws.protection.allowFormatColumns = False
                ws.protection.allowFormatRows = False
                ws.protection.allowSort = False
                ws.protection.allowAutoFilter = False
            
            wb.save(caminho_arquivo)
        except Exception as e:
            print(f"⚠️ Aviso: Erro ao proteger planilhas: {e}")

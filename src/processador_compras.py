"""
Sistema de Processamento Automático de Compras
Monitora cliques em "COMPRAR" nos Excels e processa automaticamente
"""
import threading
import time
import json
import os
from typing import Dict, Optional, Tuple
from datetime import datetime
from pathlib import Path
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - COMPRAS - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ProcessadorCompras:
    """Processa compras automaticamente quando cliques são detectados"""
    
    def __init__(self, api, pasta_monitoramento: Optional[str] = None):
        """
        Inicializa o processador de compras
        
        Args:
            api: Instância de APIGranpix
            pasta_monitoramento: Pasta para armazenar logs de compras
        """
        self.api = api
        self.pasta_monitoramento = pasta_monitoramento or "data/compras"
        os.makedirs(self.pasta_monitoramento, exist_ok=True)
        
        # Arquivo de log de compras
        self.arquivo_log = os.path.join(self.pasta_monitoramento, "historico_compras.json")
        
        # Thread de monitoramento
        self.rodando = False
        self.thread_monitor: Optional[threading.Thread] = None
        self.lock = threading.Lock()
        
        logger.info("🛒 Processador de Compras inicializado")
    
    def processar_compra_carro(self, equipe_id: str, modelo_id: str) -> Tuple[bool, str]:
        """
        Processa compra de um carro
        
        Args:
            equipe_id: ID da equipe
            modelo_id: ID do modelo de carro
            
        Returns:
            (sucesso, mensagem)
        """
        # Tentar obter do gerenciador ou direto do banco
        equipe = self.api.obter_info_equipe(equipe_id)
        if not equipe:
            equipe = self.api.db.carregar_equipe(equipe_id)
        if not equipe:
            return False, f"❌ Equipe não encontrada: {equipe_id}"
        
        # Obter modelo
        modelo = None
        if self.api.loja_carros:
            for m in self.api.loja_carros.modelos:
                if m.id == modelo_id:
                    modelo = m
                    break
        
        if not modelo:
            return False, f"❌ Modelo não encontrado: {modelo_id}"
        
        # Verificar saldo
        if equipe.doricoins < modelo.preco:
            return False, f"❌ Saldo insuficiente. Precisa: 💰 {modelo.preco:,.2f} | Tem: 💰 {equipe.doricoins:,.2f}"
        
        try:
            # Comprar carro (o método api.comprar_carro já desconta o saldo e salva)
            resultado = self.api.comprar_carro(equipe_id, modelo_id)
            
            if not resultado:
                return False, f"❌ Erro ao comprar carro {modelo.modelo}"
            
            # Registrar compra
            self._registrar_compra(equipe_id, "CARRO", modelo.modelo, modelo.preco)
            
            # Disparar auto-export
            if self.api.auto_export_monitor:
                self.api.auto_export_monitor.registrar_mudanca(equipe_id)
            
            msg = f"✅ Carro {modelo.marca} {modelo.modelo} comprado por 💰 {modelo.preco:,.2f}"
            logger.info(f"COMPRA CARRO: {equipe.nome} - {msg}")
            return True, msg
            
        except Exception as e:
            logger.error(f"Erro ao processar compra de carro: {e}")
            return False, f"❌ Erro: {str(e)}"
    
    def processar_compra_peca(self, equipe_id: str, peca_id: str) -> Tuple[bool, str]:
        """
        Processa compra de uma peça
        
        Args:
            equipe_id: ID da equipe
            peca_id: ID da peça
            
        Returns:
            (sucesso, mensagem)
        """
        # Tentar obter do gerenciador ou direto do banco
        equipe = self.api.obter_info_equipe(equipe_id)
        if not equipe:
            equipe = self.api.db.carregar_equipe(equipe_id)
        if not equipe:
            return False, f"❌ Equipe não encontrada: {equipe_id}"
        
        # Obter peça
        peca = None
        if self.api.loja_pecas:
            for p in self.api.loja_pecas.pecas:
                if p.id == peca_id:
                    peca = p
                    break
        
        if not peca:
            return False, f"❌ Peça não encontrada: {peca_id}"
        
        # Verificar saldo
        if equipe.doricoins < peca.preco:
            return False, f"❌ Saldo insuficiente. Precisa: 💰 {peca.preco:,.2f} | Tem: 💰 {equipe.doricoins:,.2f}"
        
        # Verificar se tem carro
        if not equipe.carro:
            return False, "❌ Equipe não tem carro!"
        
        # Verificar compatibilidade da peça com o carro
        if peca.compatibilidade != "universal" and peca.compatibilidade != equipe.carro.modelo_id:
            modelo_carro = self.api.db.carregar_modelo_por_id(equipe.carro.modelo_id)
            modelo_peca = self.api.db.carregar_modelo_por_id(peca.compatibilidade)
            return False, (
                f"❌ A peça {peca.nome} não é compatível com seu carro {modelo_carro.marca} {modelo_carro.modelo}. "
                f"Esta peça é específica para {modelo_peca.marca} {modelo_peca.modelo}."
            )
        
        try:
            # Comprar peça (o método api.comprar_peca já desconta o saldo e salva)
            resultado = self.api.comprar_peca(equipe_id, peca_id)
            
            if not resultado:
                return False, f"❌ Erro ao comprar peça {peca.nome}"
            
            # Registrar compra
            self._registrar_compra(equipe_id, "PEÇA", peca.nome, peca.preco)
            
            # Disparar auto-export
            if self.api.auto_export_monitor:
                self.api.auto_export_monitor.registrar_mudanca(equipe_id)
            
            msg = f"✅ Peça {peca.nome} instalada por 💰 {peca.preco:,.2f}"
            logger.info(f"COMPRA PEÇA: {equipe.nome} - {msg}")
            return True, msg
            
        except Exception as e:
            logger.error(f"Erro ao processar compra de peça: {e}")
            return False, f"❌ Erro: {str(e)}"
    
    def _registrar_compra(self, equipe_id: str, tipo: str, item: str, valor: float):
        """Registra compra no arquivo de histórico
        
        Args:
            equipe_id: ID da equipe
            tipo: Tipo (CARRO ou PEÇA)
            item: Nome do item
            valor: Preço pago
        """
        try:
            # Carregar histórico
            historico = []
            if os.path.exists(self.arquivo_log):
                with open(self.arquivo_log, 'r', encoding='utf-8') as f:
                    historico = json.load(f)
            
            # Adicionar nova compra
            compra = {
                'timestamp': datetime.now().isoformat(),
                'equipe_id': equipe_id,
                'tipo': tipo,
                'item': item,
                'valor': valor
            }
            historico.append(compra)
            
            # Salvar histórico
            with open(self.arquivo_log, 'w', encoding='utf-8') as f:
                json.dump(historico, f, indent=2, ensure_ascii=False)
        except Exception as e:
            logger.error(f"Erro ao registrar compra: {e}")
    
    def obter_historico_compras(self, equipe_id: Optional[str] = None):
        """Retorna histórico de compras
        
        Args:
            equipe_id: Filtrar por equipe (opcional)
            
        Returns:
            Lista de compras
        """
        try:
            if not os.path.exists(self.arquivo_log):
                return []
            
            with open(self.arquivo_log, 'r', encoding='utf-8') as f:
                historico = json.load(f)
            
            if equipe_id:
                return [c for c in historico if c['equipe_id'] == equipe_id]
            
            return historico
        except Exception as e:
            logger.error(f"Erro ao carregar histórico: {e}")
            return []
    
    def obter_status(self) -> dict:
        """Retorna status do processador"""
        return {
            'rodando': self.rodando,
            'pasta_monitoramento': self.pasta_monitoramento,
            'arquivo_log': self.arquivo_log,
            'total_compras': len(self.obter_historico_compras())
        }


class MonitorComprasExcel:
    """Monitora mudanças em Excels para detectar compras"""
    
    def __init__(self, processador: ProcessadorCompras, pasta_equipes: str):
        """
        Inicializa monitor de compras em Excel
        
        Args:
            processador: Instância de ProcessadorCompras
            pasta_equipes: Pasta onde os Excels são salvos
        """
        self.processador = processador
        self.pasta_equipes = Path(pasta_equipes)
        self.rodando = False
        self.thread_monitor: Optional[threading.Thread] = None
        self.lock = threading.Lock()
        
        # Rastrear últimas modificações
        self.ultimas_modificacoes: Dict[str, float] = {}
        
        logger.info("👀 Monitor de Compras Excel inicializado")
    
    def iniciar(self):
        """Inicia o monitoramento"""
        if self.rodando:
            return
        
        self.rodando = True
        self.thread_monitor = threading.Thread(
            target=self._monitor_loop,
            daemon=True,
            name="MonitorComprasExcel"
        )
        self.thread_monitor.start()
        logger.info("🚀 Monitor de compras iniciado")
    
    def parar(self):
        """Para o monitoramento"""
        self.rodando = False
        if self.thread_monitor:
            self.thread_monitor.join(timeout=5)
        logger.info("⛔ Monitor de compras parado")
    
    def _monitor_loop(self):
        """Loop principal de monitoramento"""
        while self.rodando:
            try:
                self._verificar_mudancas()
                time.sleep(3)  # Verificar a cada 3 segundos
            except Exception as e:
                logger.error(f"Erro no loop de monitoramento: {e}")
                time.sleep(1)
    
    def _verificar_mudancas(self):
        """Verifica mudanças nos arquivos Excel"""
        if not self.pasta_equipes.exists():
            return
        
        for arquivo in self.pasta_equipes.glob("*.xlsx"):
            try:
                stat = arquivo.stat()
                tempo_modificacao = stat.st_mtime
                
                # Verificar se arquivo foi modificado
                tempo_anterior = self.ultimas_modificacoes.get(arquivo.name, 0)
                
                if tempo_modificacao > tempo_anterior and tempo_modificacao > (time.time() - 5):
                    # Arquivo foi modificado nos últimos 5 segundos
                    self.ultimas_modificacoes[arquivo.name] = tempo_modificacao
                    self._processar_arquivo(arquivo)
                
                # Atualizar timestamp
                self.ultimas_modificacoes[arquivo.name] = tempo_modificacao
                
            except Exception as e:
                logger.error(f"Erro ao verificar arquivo {arquivo.name}: {e}")
    
    def _processar_arquivo(self, arquivo: Path):
        """Processa arquivo Excel para detectar compras
        
        Args:
            arquivo: Caminho do arquivo Excel
        """
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(arquivo, data_only=True)
            
            # Processar aba de Loja de Carros
            if 'Loja Carros' in wb.sheetnames:
                self._processar_loja_carros(wb['Loja Carros'], arquivo)
            
            # Processar aba de Loja de Peças
            if 'Loja Peças' in wb.sheetnames:
                self._processar_loja_pecas(wb['Loja Peças'], arquivo)
            
        except Exception as e:
            logger.error(f"Erro ao processar arquivo {arquivo.name}: {e}")
    
    def _processar_loja_carros(self, worksheet, arquivo: Path):
        """Processa cliques em Loja de Carros
        
        Args:
            worksheet: Worksheet de Loja de Carros
            arquivo: Caminho do arquivo Excel
        """
        # Extrair equipe_id do nome do arquivo
        equipe_id = self._extrair_equipe_id(arquivo)
        if not equipe_id:
            return
        
        # Verificar coluna de ação (G)
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), 2):
            if len(row) >= 7:
                cell_acao = row[6]  # Coluna G
                if cell_acao.value == "COMPRAR":
                    # Extrair ID do modelo (coluna A)
                    cell_id = row[0]
                    modelo_id = cell_id.value
                    
                    if modelo_id:
                        logger.info(f"🛒 Compra detectada: {equipe_id} → Carro {modelo_id}")
                        sucesso, msg = self.processador.processar_compra_carro(equipe_id, modelo_id)
                        logger.info(msg)
    
    def _processar_loja_pecas(self, worksheet, arquivo: Path):
        """Processa cliques em Loja de Peças
        
        Args:
            worksheet: Worksheet de Loja de Peças
            arquivo: Caminho do arquivo Excel
        """
        # Extrair equipe_id do nome do arquivo
        equipe_id = self._extrair_equipe_id(arquivo)
        if not equipe_id:
            return
        
        # Verificar coluna de ação (F)
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), 2):
            if len(row) >= 6:
                cell_acao = row[5]  # Coluna F
                if cell_acao.value == "COMPRAR":
                    # Extrair ID da peça (coluna A)
                    cell_id = row[0]
                    peca_id = cell_id.value
                    
                    if peca_id:
                        logger.info(f"🛒 Compra detectada: {equipe_id} → Peça {peca_id}")
                        sucesso, msg = self.processador.processar_compra_peca(equipe_id, peca_id)
                        logger.info(msg)
    
    def _extrair_equipe_id(self, arquivo: Path) -> Optional[str]:
        """Extrai equipe_id do nome do arquivo ou conteúdo
        
        Args:
            arquivo: Caminho do arquivo Excel
            
        Returns:
            ID da equipe ou None
        """
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(arquivo, data_only=True)
            
            # Tentar obter da aba 'Equipe'
            if 'Equipe' in wb.sheetnames:
                ws = wb['Equipe']
                # Procurar pela linha "ID da Equipe"
                for row in ws.iter_rows(values_only=True, min_row=2, max_row=10):
                    if row[0] == 'ID da Equipe' and len(row) > 1:
                        return row[1]
            
            return None
        except Exception as e:
            logger.error(f"Erro ao extrair equipe_id: {e}")
            return None

"""
Exportador de equipes para Excel com formatação profissional
"""
import os
from datetime import datetime
from typing import Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import Equipe, TipoPeca


class ExportadorExcelProfissional:
    """Exporta dados de equipes para Excel com formatação profissional"""
    
    def __init__(self, pasta_saida: str = None):
        """Inicializa exportador
        
        Args:
            pasta_saida: Pasta onde salvar os arquivos
                         Se None, usa OneDrive padrão
        """
        if pasta_saida is None:
            home = os.path.expanduser("~")
            pasta_saida = os.path.join(home, "OneDrive", "GRANPIX", "equipes")
        
        self.pasta_saida = pasta_saida
        os.makedirs(self.pasta_saida, exist_ok=True)
    
    def exportar_equipe(self, equipe: Equipe) -> str:
        """Exporta equipe para Excel formatado
        
        Args:
            equipe: Equipe a exportar
            
        Returns:
            Caminho do arquivo gerado
        """
        # Nome do arquivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"{equipe.nome.replace(' ', '_').lower()}_{timestamp}.xlsx"
        caminho_arquivo = os.path.join(self.pasta_saida, nome_arquivo)
        
        # Criar writer Excel
        with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
            # Aba 1: Resumo da Equipe
            self._criar_aba_resumo(equipe, writer)
            
            # Aba 2: Peças do Carro
            self._criar_aba_pecas(equipe, writer)
            
            # Aba 3: Histórico de Compras
            self._criar_aba_historico(equipe, writer)
            
            # Aba 4: Pilotos e Estatísticas
            self._criar_aba_pilotos(equipe, writer)
            
            # Aba 5: Dashboard Financeiro
            self._criar_aba_financeiro(equipe, writer)
        
        # Aplicar formatação
        self._aplicar_formatacao(caminho_arquivo)
        
        return caminho_arquivo
    
    def _criar_aba_resumo(self, equipe: Equipe, writer):
        """Cria aba de resumo da equipe"""
        dados = {
            'Propriedade': [
                'Nome da Equipe',
                'ID',
                'Data de Criação',
                'Saldo Atual',
                'Total de Pilotos',
                'Total de Peças',
                'Saúde Média Peças',
                'Total de Batalhas',
                'Vitórias',
                'Derrotas',
                'Empates'
            ],
            'Valor': [
                equipe.nome,
                equipe.id,
                equipe.data_criacao.strftime("%d/%m/%Y %H:%M"),
                f"💰 {equipe.saldo:,.2f}",
                len(equipe.pilotos),
                len(equipe.pecas),
                f"{equipe.get_saude_media_pecas():.1f}%",
                equipe.get_vitoria_total() + equipe.get_derrota_total() + equipe.get_empate_total(),
                f"✓ {equipe.get_vitoria_total()}",
                f"✗ {equipe.get_derrota_total()}",
                f"⚖ {equipe.get_empate_total()}"
            ]
        }
        
        df = pd.DataFrame(dados)
        df.to_excel(writer, sheet_name='Resumo', index=False)
    
    def _criar_aba_pecas(self, equipe: Equipe, writer):
        """Cria aba de peças do carro"""
        if not equipe.pecas:
            df = pd.DataFrame({'Aviso': ['Nenhuma peça registrada']})
            df.to_excel(writer, sheet_name='Peças Carro', index=False)
            return
        
        dados = {
            'Peça': [],
            'Tipo': [],
            'Saúde': [],
            'Status': [],
            'Preço': [],
            'Data Compra': []
        }
        
        for peca in equipe.pecas:
            dados['Peça'].append(peca.nome)
            dados['Tipo'].append(peca.tipo.value.upper())
            dados['Saúde'].append(f"{peca.saude:.1f}%")
            dados['Status'].append(peca.get_status())
            dados['Preço'].append(f"💰 {peca.preco:,.2f}")
            dados['Data Compra'].append(peca.data_compra.strftime("%d/%m/%Y"))
        
        df = pd.DataFrame(dados)
        df.to_excel(writer, sheet_name='Peças Carro', index=False)
    
    def _criar_aba_historico(self, equipe: Equipe, writer):
        """Cria aba de histórico de compras"""
        if not equipe.historico_compras:
            df = pd.DataFrame({'Aviso': ['Nenhuma transação registrada']})
            df.to_excel(writer, sheet_name='Histórico Compras', index=False)
            return
        
        dados = {
            'Data': [],
            'Tipo': [],
            'Descrição': [],
            'Valor': [],
            'Saldo Anterior': [],
            'Saldo Posterior': []
        }
        
        for compra in equipe.historico_compras:
            dados['Data'].append(compra.data.strftime("%d/%m/%Y %H:%M"))
            dados['Tipo'].append(compra.tipo.value)
            dados['Descrição'].append(compra.descricao)
            dados['Valor'].append(f"💰 {compra.valor:,.2f}")
            dados['Saldo Anterior'].append(f"💰 {compra.saldo_anterior:,.2f}")
            dados['Saldo Posterior'].append(f"💰 {compra.saldo_posterior:,.2f}")
        
        df = pd.DataFrame(dados)
        df.to_excel(writer, sheet_name='Histórico Compras', index=False)
    
    def _criar_aba_pilotos(self, equipe: Equipe, writer):
        """Cria aba de pilotos e estatísticas"""
        if not equipe.pilotos:
            df = pd.DataFrame({'Aviso': ['Nenhum piloto registrado']})
            df.to_excel(writer, sheet_name='Pilotos', index=False)
            return
        
        dados = {
            'Piloto': [],
            'Vitórias': [],
            'Derrotas': [],
            'Empates': [],
            'Total Batalhas': [],
            'Taxa Vitória': []
        }
        
        for piloto in equipe.pilotos:
            total = piloto.vitoria + piloto.derrota + piloto.empate
            taxa = piloto.get_taxa_vitoria()
            
            dados['Piloto'].append(piloto.nome)
            dados['Vitórias'].append(f"✓ {piloto.vitoria}")
            dados['Derrotas'].append(f"✗ {piloto.derrota}")
            dados['Empates'].append(f"⚖ {piloto.empate}")
            dados['Total Batalhas'].append(total)
            dados['Taxa Vitória'].append(f"{taxa:.1f}%")
        
        df = pd.DataFrame(dados)
        df.to_excel(writer, sheet_name='Pilotos', index=False)
    
    def _criar_aba_financeiro(self, equipe: Equipe, writer):
        """Cria aba de resumo financeiro"""
        # Calcular totais
        total_compras = sum(c.valor for c in equipe.historico_compras if c.tipo.value == "Compra")
        total_premios = sum(c.valor for c in equipe.historico_compras if c.tipo.value in ["Prêmio Vitória", "Salário"])
        total_vendas = sum(c.valor for c in equipe.historico_compras if c.tipo.value == "Venda")
        
        dados = {
            'Métrica Financeira': [
                'Saldo Inicial (Estimado)',
                'Total de Compras',
                'Total de Vendas',
                'Total de Prêmios/Salários',
                'Saldo Atual',
                '',
                'Número de Transações',
                'Ticket Médio Compra'
            ],
            'Valor': [
                f"💰 {equipe.saldo + total_compras - total_premios - total_vendas:,.2f}",
                f"💰 -{total_compras:,.2f}",
                f"💰 +{total_vendas:,.2f}",
                f"💰 +{total_premios:,.2f}",
                f"💰 {equipe.saldo:,.2f}",
                '',
                len(equipe.historico_compras),
                f"💰 {total_compras / max(1, len([c for c in equipe.historico_compras if c.tipo.value == 'Compra'])):,.2f}"
            ]
        }
        
        df = pd.DataFrame(dados)
        df.to_excel(writer, sheet_name='Financeiro', index=False)
    
    def _aplicar_formatacao(self, caminho_arquivo: str):
        """Aplica formatação profissional ao Excel"""
        wb = load_workbook(caminho_arquivo)
        
        # Definir cores e estilos
        cor_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cor_header_text = Font(bold=True, color="FFFFFF", size=12)
        
        cor_alternada = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        borda = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        alinhamento_centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
        alinhamento_esquerda = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Aplicar formatação para cada aba
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Dimensionar colunas
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Formatar header (primeira linha)
            for cell in ws[1]:
                cell.fill = cor_header
                cell.font = cor_header_text
                cell.alignment = alinhamento_centro
                cell.border = borda
            
            # Formatar linhas de dados
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                cor_fundo = cor_alternada if row_idx % 2 == 0 else PatternFill()
                
                for cell in row:
                    cell.fill = cor_fundo
                    cell.border = borda
                    
                    # Alinhamento baseado no tipo
                    if any(keyword in str(cell.value) for keyword in ['💰', '%', '✓', '✗', '⚖']):
                        cell.alignment = alinhamento_centro
                    else:
                        cell.alignment = alinhamento_esquerda
            
            # Congelar primeira linha
            ws.freeze_panes = 'A2'
        
        wb.save(caminho_arquivo)
    
    def exportar_todas_equipes(self, equipes: list) -> list:
        """Exporta múltiplas equipes"""
        arquivos = []
        for equipe in equipes:
            try:
                caminho = self.exportar_equipe(equipe)
                arquivos.append(caminho)
                print(f"✓ {equipe.nome} exportado")
            except Exception as e:
                print(f"✗ Erro ao exportar {equipe.nome}: {e}")
        
        return arquivos
